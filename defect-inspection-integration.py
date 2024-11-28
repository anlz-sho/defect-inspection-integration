# モジュールインポート
import os
import time
import openpyxl as excel

from openpyxl.utils.exceptions import InvalidFileException  # openpyxlの例外処理のため

##############################
# 関数定義
def active_wf():    # 書込中のwfをコマンドラインに表示
    lot_name = fn[10:21]
    wf_name = fn[22:24]
    print(lot_name + '_' + wf_name)

def sheet_name():   # シート名を変更
    lot_name = fn[10:21]
    wf_name = fn[22:24]
    xs.title = lot_name + '_' + wf_name

def xlcopy():       # エクセルファイルのコピー
    for i in range(1, rw + 1):
        ws_a = ws.cell(row=i, column=1).value   # A列の値をコピー
        ws_b = ws.cell(row=i, column=2).value   # B列の値をコピー
        ws_h = ws.cell(row=i, column=7).value   # G列の値をコピー
        ws_i = ws.cell(row=i, column=8).value   # H列の値をコピー
        ws_j = ws.cell(row=i, column=9).value   # I列の値をコピー
        ws_k = ws.cell(row=i, column=10).value  # J列の値をコピー
        ws_m = ws.cell(row=i, column=12).value  # L列の値をコピー
        ws_n = ws.cell(row=i, column=13).value  # M列の値をコピー
        ws_o = ws.cell(row=i, column=14).value  # N列の値をコピー
        ws_p = ws.cell(row=i, column=15).value  # O列の値をコピー
        ws_r = ws.cell(row=i, column=17).value  # Q列の値をコピー
        ws_s = ws.cell(row=i, column=18).value  # R列の値をコピー
        ws_t = ws.cell(row=i, column=19).value  # S列の値をコピー

        xs.cell(row=i, column=1).value = ws_a   # A列へ貼付
        xs.cell(row=i, column=2).value = ws_b   # B列へ貼付
        xs.cell(row=i, column=7).value = ws_h   # G列へ貼付
        xs.cell(row=i, column=8).value = ws_i   # H列へ貼付
        xs.cell(row=i, column=9).value = ws_i   # I列へ貼付
        xs.cell(row=i, column=10).value = ws_j  # J列へ貼付
        xs.cell(row=i, column=12).value = ws_k  # L列へ貼付
        xs.cell(row=i, column=13).value = ws_m  # M列へ貼付
        xs.cell(row=i, column=14).value = ws_n  # N列へ貼付
        xs.cell(row=i, column=15).value = ws_o  # O列へ貼付
        xs.cell(row=i, column=17).value = ws_p  # Q列へ貼付
        xs.cell(row=i, column=18).value = ws_r  # R列へ貼付
        xs.cell(row=i, column=19).value = ws_s  # S列へ貼付


##############################
# パス設定
path = "/path/to/sample/data"           # 目視結果エクセル置き場
xpath = "/path/to/sample/"              # まとめ先エクセル置き場
output = 'output.xlsx'                  # まとめ先エクセルファイル名

##############################
# タイマースタート
st = time.time()
xw = excel.load_workbook(os.path.join(xpath, output))

##############################

for fn in os.listdir(path):                             # 目視結果エクセルをコピー
    try:
        wb = excel.load_workbook(os.path.join(path, fn))
        active_wf()
        ws = wb['log']
        xs = xw.create_sheet()
        sheet_name()

        for k in reversed(range(1, ws.max_row + 1)):    # 目視漏れの検出用
            if ws.cell(row=k, column=1).value is not None:
                rw = k
                if ws.max_row != rw:
                    print("目視ミス")
                    print("A列は", k, "行目まで記載有り")
                    print(ws.max_row, "行目にミスあり")
                break

        xlcopy()
        wb.close()

    except InvalidFileException:
        continue

xw.save(os.path.join(xpath, output))                    # 保存
process_time = time.time() - st
print('Process Time:' + str(process_time))




